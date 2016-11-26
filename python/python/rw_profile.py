import re
import pdb
import linecache 
from openpyxl.writer.excel import ExcelWriter 
from openpyxl.workbook import Workbook

#打开文件
infilename = '/home/zl/git/code/python/profile_big.txt'
f=open(infilename, 'r')
wb=Workbook()
sheet = wb.active

#输入文件标题行号
inFileTitleLine = 3
#时间行号
timeLine = 2
#标题行
titleLine = 3
#数据起始行
dataLineStart = 4
#数据计数
dataLineCount = 0
#数据记录行
dataRecords = []
#默认默认起始列
dataColumnStartDefault = 1
#起始列
dataColumnStart = dataColumnStartDefault
#excel列数
columnCount = 1
#目标列标题(提取的列)
targetTitle = ['Chunk','v_vy','v_vz']
#时间步长
timeStep = 10000

#所有标题列表
titleList = []

#调试信息
# pdb.set_trace()

print ('开始')
print ('开始解析标题')
titleRecord = linecache.getline(infilename,inFileTitleLine)

#说明是标题行
#去掉行的空格
titleRecord = titleRecord.strip()
titleList = re.compile("\s+").split(titleRecord)
titleList.remove('#')

#该步长的数据行是否解析
isThisStepDataBeAnalyze = True
print('循环解析数据行')
for line in f.readlines():
    if(line.find('#')==0):
        continue
    #读取数据行
    line = line.strip()
    dataRecords = re.compile("\s+").split(line)
    #调试信息
    # pdb.set_trace()
    if len(dataRecords) == 3:
        #步长判断
        if int(dataRecords[0])%timeStep != 0:
            isThisStepDataBeAnalyze = False
            continue
        else:
            isThisStepDataBeAnalyze = True
        #调试信息
        # pdb.set_trace()
        #重新赋值起始行号
        dataLineCount = dataLineStart

        if dataColumnStart == dataColumnStartDefault:
            dataColumnStartDefault = 0
        else:
            dataColumnStart += len(targetTitle)
        
        #赋值当前列的标志
        titleColumnNum = dataColumnStart
        #打印时间步
        sheet.cell(row =timeLine, column = dataColumnStart).value = dataRecords[0]
        #打印标题
        for titleName in targetTitle:
            sheet.cell(row =titleLine, column = titleColumnNum).value = titleName
            titleColumnNum += 1
        continue
    else:
        if isThisStepDataBeAnalyze == False:
            #该数据行不被解析继续循环
            continue
        #打印数据行
        for title in targetTitle:
            #标题在所有标题list中的位置
            titleIndex = titleList.index(title)
            #标题在目标list中的位置
            titleTargetIndex = targetTitle.index(title)
            #直接将对应的值写入到excel
            sheet.cell(row = dataLineCount,column = dataColumnStart + titleTargetIndex).value = dataRecords[titleIndex]
    #excel行号+1   
    dataLineCount += 1
    #excel　最大行列数　1048576 * 16384 在超过该值时系统自动退出
    if dataColumnStart > 16300 or dataLineCount > 1048500:
        print ("行或列超出excel 范围,保存已经解析的值并退出")
        break
    
wb.save(filename = '/home/zl/Desktop/profile.xlsx')
